"""
One-time fix: Update Jiaxi FAs that were imported with spectral_reflectance_requirement='alpha'
but should be 'Visible Spectrum Only'.

Usage:
  python manage.py shell < scripts/fix_jiaxi_spectral.py
"""
import openpyxl
from inspections.models import FirstArticleInspection

wb = openpyxl.load_workbook('sheetData/Jiaxi - First Article Workbook.xlsx', data_only=True)
ws = wb['FirstArticleSummary']

seq = 1
updated = 0
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    status = row[0]
    if not status:
        continue
    fa_id = f'JXI-FA-{seq:04d}'
    spectral_raw = str(row[5] or '').strip().lower()
    if 'visible' in spectral_raw:
        n = FirstArticleInspection.objects.filter(
            fai_id=fa_id, spectral_reflectance_requirement='alpha'
        ).update(spectral_reflectance_requirement='Visible Spectrum Only')
        if n:
            print(f'  Updated {fa_id}')
            updated += n
    seq += 1

print(f'\nDone. Updated {updated} FAs.')
