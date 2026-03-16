"""
Import FA data from a partner Excel workbook (.xlsx).

Usage:
    python manage.py import_workbook "sheetData/Milliken - First Article Workbook.xlsx" --partner-code MIL --partner-name "Milliken" --standard
    python manage.py import_workbook "path/to/workbook.xlsx" --partner-code MIL --partner-name "Milliken" --standard --dry-run
"""
import openpyxl
from datetime import date, datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.contrib.auth.models import User
from django.utils import timezone

from accounts.models import PartnerCompany
from inspections.models import (
    FirstArticleInspection, FAEvaluation, FAColorEvaluation,
    SHADE_RATING_CHOICES,
)
from core.models import CamouflageType

HISTORIC_SENTINEL = 'N/A - Historic First Article'
PLACEHOLDER_DATE = date(2000, 1, 1)

SKIP_SHEETS = frozenset({
    'FirstArticleSummary', 'DATA',
    'MC', 'MCAlpine', 'MCTropic', 'MCBlack', 'MCArid',
})

VALID_RATINGS = frozenset(code for code, _ in SHADE_RATING_CHOICES)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def to_str(value, max_length=None):
    if value is None or value == '':
        return ''
    if isinstance(value, float):
        return str(int(value)) if value == int(value) else str(value)
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d')
    s = str(value).strip()
    return s[:max_length] if max_length else s


def to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    s = str(value).strip()
    if s == HISTORIC_SENTINEL or not s:
        return None
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def to_datetime_aware(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return timezone.make_aware(value) if timezone.is_naive(value) else value
    if isinstance(value, date):
        dt = datetime.combine(value, datetime.min.time())
        return timezone.make_aware(dt)
    s = str(value).strip()
    if s == HISTORIC_SENTINEL or not s:
        return None
    for fmt in ('%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            dt = datetime.strptime(s, fmt)
            return timezone.make_aware(dt)
        except ValueError:
            continue
    return None


def is_historic_value(value):
    return value is None or str(value).strip() == HISTORIC_SENTINEL


def parse_rating_code(rating_str):
    if not rating_str or str(rating_str).strip() in ('', 'Select', 'Pending'):
        return ''
    code = str(rating_str).strip().split(' - ', 1)[0].strip()
    return code if code in VALID_RATINGS else ''


def parse_pass_fail(value):
    if not value:
        return ''
    v = str(value).strip().lower()
    return v if v in ('pass', 'fail') else ''


def map_variant_name(raw_name):
    if not raw_name:
        return None
    low = str(raw_name).lower().replace('\u00ae', '').replace('\ufffd', '').strip()
    if 'alpine' in low:
        return 'Multicam Alpine'
    if 'tropic' in low:
        return 'Multicam Tropic'
    if 'black' in low:
        return 'Multicam Black'
    if 'arid' in low:
        return 'Multicam Arid'
    if 'multicam' in low:
        return 'Multicam'
    return None


def parse_submitter_name(name_str):
    if not name_str or is_historic_value(name_str):
        return '', ''
    parts = str(name_str).strip().split(',', 1)
    last_name = parts[0].strip()
    first_name = parts[1].strip() if len(parts) > 1 else ''
    return first_name, last_name


def find_eval_sheet(generated_name, sheet_names):
    if not generated_name:
        return None
    g = str(generated_name).strip()
    if g in sheet_names:
        return g
    for name in sheet_names:
        if g.startswith(name) or name.startswith(g):
            return name
    return None


def parse_eval_sheet(ws):
    colors = []
    for row_num in range(7, 14):
        color_name = ws.cell(row=row_num, column=2).value
        if not color_name:
            continue
        colors.append({
            'color_name': str(color_name).strip(),
            'rating': parse_rating_code(ws.cell(row=row_num, column=3).value),
            'comment': to_str(ws.cell(row=row_num, column=4).value, 200),
        })

    primary_raw = ws.cell(row=18, column=1).value
    primary_result = None
    if primary_raw:
        v = str(primary_raw).strip().lower()
        if v == 'fail':
            primary_result = 'fail'
        elif v in ('pass', 'sent to crye', 'pending'):
            primary_result = 'pass'

    final_raw = ws.cell(row=20, column=1).value
    final_result = None
    if final_raw:
        v = str(final_raw).strip().lower()
        if v in ('pass', 'fail'):
            final_result = v

    return {
        'colors': colors,
        'pattern_execution': parse_pass_fail(ws.cell(row=14, column=3).value),
        'pattern_execution_comment': to_str(ws.cell(row=14, column=4).value, 500),
        'scale': parse_pass_fail(ws.cell(row=15, column=3).value),
        'scale_comment': to_str(ws.cell(row=15, column=4).value, 500),
        'spectral_reflectance': parse_pass_fail(ws.cell(row=16, column=3).value),
        'spectral_reflectance_comment': to_str(ws.cell(row=16, column=4).value, 500),
        'primary_result': primary_result,
        'primary_comments': to_str(ws.cell(row=18, column=3).value),
        'final_result': final_result,
        'final_comments': to_str(ws.cell(row=20, column=3).value),
    }


# ---------------------------------------------------------------------------
# Command
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = 'Import FA data from a partner Excel workbook'

    def add_arguments(self, parser):
        parser.add_argument('workbook', help='Path to the Excel workbook')
        parser.add_argument('--partner-code', required=True, help='Company code (e.g. MIL)')
        parser.add_argument('--partner-name', required=True, help='Company name (e.g. Milliken)')
        parser.add_argument('--standard', action='store_true', help='Mark as standard partner')
        parser.add_argument('--narrow', action='store_true', help='Mark as narrow partner')
        parser.add_argument('--dry-run', action='store_true', help='Preview without saving')

    def handle(self, *args, **options):
        wb_path = options['workbook']
        partner_code = options['partner_code'].upper()
        partner_name = options['partner_name']
        dry_run = options['dry_run']

        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN — no changes will be made\n'))

        try:
            wb = openpyxl.load_workbook(wb_path, data_only=True)
        except Exception as e:
            raise CommandError(f'Cannot open workbook: {e}')

        if 'FirstArticleSummary' not in wb.sheetnames:
            raise CommandError('Workbook missing FirstArticleSummary sheet')

        eval_sheet_names = [n for n in wb.sheetnames if n not in SKIP_SHEETS]
        self.stdout.write(f'Evaluation sheets found: {len(eval_sheet_names)}')

        # Inspector accounts (only required for real imports)
        primary_user = None
        final_user = None
        if not dry_run:
            try:
                primary_user = User.objects.get(username='primary_inspector')
            except User.DoesNotExist:
                raise CommandError('User "primary_inspector" not found.')
            try:
                final_user = User.objects.get(username='final_inspector')
            except User.DoesNotExist:
                raise CommandError('User "final_inspector" not found.')

        # Camouflage type cache: { db_name: { obj, colors: {color_name: VariantColor} } }
        camo_cache = {}
        if not dry_run:
            for ct in CamouflageType.objects.prefetch_related('colors'):
                camo_cache[ct.camouflage_name] = {
                    'obj': ct,
                    'colors': {c.color_name: c for c in ct.colors.all()},
                }
            if not camo_cache:
                raise CommandError('No CamouflageType records. Run load_initial_data first.')

        # Partner company
        company = None
        if not dry_run:
            company, created = PartnerCompany.objects.get_or_create(
                code=partner_code,
                defaults={
                    'name': partner_name,
                    'is_standard': options['standard'],
                    'is_narrow': options['narrow'],
                    'status': 'active',
                },
            )
            action = 'Created' if created else 'Using existing'
            self.stdout.write(self.style.SUCCESS(f'{action} partner: {company.name} ({company.code})'))
        else:
            self.stdout.write(f'[DRY RUN] Partner: {partner_name} ({partner_code})')

        # Parse summary and import
        stats = dict(
            fa_created=0, fa_historic=0, fa_with_eval=0,
            primary_evals=0, final_evals=0, color_evals=0,
            errors=0, skipped=0,
        )

        ws = wb['FirstArticleSummary']
        row_idx = 3
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            row_idx += 1
            status_raw = row[0]
            if not status_raw or str(status_raw).strip() == '':
                continue

            status = str(status_raw).strip().lower()
            if status not in ('approved', 'rejected', 'pending'):
                self.stdout.write(self.style.WARNING(
                    f'  Row {row_idx}: Unknown status "{status_raw}", skipping'
                ))
                stats['skipped'] += 1
                continue

            fabric_style = to_str(row[1], 200)
            variant_name = map_variant_name(row[2])
            if not variant_name:
                self.stdout.write(self.style.ERROR(
                    f'  Row {row_idx}: Unknown variant "{row[2]}", skipping'
                ))
                stats['errors'] += 1
                continue

            if not dry_run:
                if variant_name not in camo_cache:
                    self.stdout.write(self.style.ERROR(
                        f'  Row {row_idx}: Variant "{variant_name}" not in DB, skipping'
                    ))
                    stats['errors'] += 1
                    continue
                camo_data = camo_cache[variant_name]
                multicam_variant = camo_data['obj']
                variant_colors = camo_data['colors']
            else:
                multicam_variant = None
                variant_colors = {}

            is_historic = is_historic_value(row[7])

            shade_standard = str(row[3] or '').strip().lower()
            if shade_standard not in ('alpha', 'beta'):
                shade_standard = 'alpha'

            shade_std_num = '' if is_historic_value(row[4]) else to_str(row[4], 20)

            spectral = str(row[5] or '').strip().lower()
            if spectral not in ('alpha', 'beta', 'swir'):
                spectral = 'alpha'

            lot_number = '' if is_historic_value(row[6]) else to_str(row[6], 50)
            if not lot_number:
                lot_number = f'HISTORIC-{row_idx}'

            date_of_printing = to_date(row[7]) or PLACEHOLDER_DATE
            ship_date = to_date(row[8])
            tracking = '' if is_historic_value(row[9]) else to_str(row[9], 50)
            first_name, last_name = parse_submitter_name(row[10])
            submission_dt = to_datetime_aware(row[12])
            sheet_name = to_str(row[13], 250)

            # --- Dry-run output ---
            if dry_run:
                tag = '[HISTORIC] ' if is_historic else ''
                self.stdout.write(
                    f'  {tag}Row {row_idx}: {status.upper()} | '
                    f'{fabric_style} | {variant_name} | Lot: {lot_number}'
                )
                if is_historic:
                    stats['fa_historic'] += 1
                else:
                    match = find_eval_sheet(sheet_name, eval_sheet_names)
                    if match:
                        self.stdout.write(f'    -> Eval sheet: {match}')
                        stats['fa_with_eval'] += 1
                    else:
                        self.stdout.write(self.style.WARNING(
                            f'    -> No eval sheet for "{sheet_name}"'
                        ))
                stats['fa_created'] += 1
                continue

            # --- Real import ---
            try:
                with transaction.atomic():
                    fa = FirstArticleInspection(
                        company=company,
                        vendor=None,
                        fabric_style=fabric_style,
                        multicam_variant=multicam_variant,
                        shade_standard=shade_standard,
                        shade_standard_number=shade_std_num,
                        spectral_reflectance_requirement=spectral,
                        fa_lot_number=lot_number,
                        date_of_printing=date_of_printing,
                        first_article_ship_date=ship_date,
                        tracking_number=tracking,
                        submitter_first_name=first_name,
                        submitter_last_name=last_name,
                        status=status,
                        submitted=True,
                        is_historic=is_historic,
                        sheet_name_generated=sheet_name,
                    )
                    fa.save()

                    if submission_dt:
                        FirstArticleInspection.objects.filter(pk=fa.pk).update(
                            submission_date=submission_dt
                        )

                    if is_historic:
                        stats['fa_historic'] += 1
                        self.stdout.write(
                            f'  [HISTORIC] {fa.fai_id}: {fabric_style} ({status})'
                        )
                    else:
                        self.stdout.write(f'  {fa.fai_id}: {fabric_style} ({status})')

                    stats['fa_created'] += 1

                    if not is_historic:
                        eval_sheet_name = find_eval_sheet(sheet_name, eval_sheet_names)
                        if eval_sheet_name:
                            eval_ws = wb[eval_sheet_name]
                            eval_data = parse_eval_sheet(eval_ws)
                            self._create_evaluations(
                                fa, eval_data, variant_colors,
                                primary_user, final_user,
                                submission_dt, stats,
                            )
                            stats['fa_with_eval'] += 1
                        else:
                            self.stdout.write(self.style.WARNING(
                                f'    No eval sheet for "{sheet_name}"'
                            ))

            except Exception as e:
                stats['errors'] += 1
                self.stdout.write(self.style.ERROR(f'  Row {row_idx}: Error — {e}'))

        # Summary
        self.stdout.write('\n' + '=' * 60)
        label = 'DRY RUN SUMMARY' if dry_run else 'IMPORT SUMMARY'
        self.stdout.write(self.style.SUCCESS(f'{label}:'))
        self.stdout.write(f'  FAs created:        {stats["fa_created"]}')
        self.stdout.write(f'    Historic:         {stats["fa_historic"]}')
        self.stdout.write(f'    With evaluations: {stats["fa_with_eval"]}')
        self.stdout.write(f'  Primary evals:      {stats["primary_evals"]}')
        self.stdout.write(f'  Final evals:        {stats["final_evals"]}')
        self.stdout.write(f'  Color evals:        {stats["color_evals"]}')
        self.stdout.write(f'  Errors:             {stats["errors"]}')
        self.stdout.write(f'  Skipped:            {stats["skipped"]}')
        self.stdout.write('=' * 60)
        wb.close()

    def _create_evaluations(self, fa, eval_data, variant_colors,
                            primary_user, final_user, submission_dt, stats):
        submitted_at = submission_dt or timezone.now()

        # ---- Primary evaluation ----
        primary_eval = FAEvaluation.objects.create(
            fa=fa,
            stage='primary',
            attempt_number=1,
            inspector=primary_user,
            pattern_execution=eval_data['pattern_execution'],
            pattern_execution_comment=eval_data['pattern_execution_comment'],
            scale=eval_data['scale'],
            scale_comment=eval_data['scale_comment'],
            spectral_reflectance=eval_data['spectral_reflectance'],
            spectral_reflectance_comment=eval_data['spectral_reflectance_comment'],
            comments=eval_data['primary_comments'],
            is_submitted=True,
            submitted_at=submitted_at,
        )
        FAEvaluation.objects.filter(pk=primary_eval.pk).update(
            evaluation_date=submitted_at
        )
        stats['primary_evals'] += 1

        for cd in eval_data['colors']:
            color_obj = variant_colors.get(cd['color_name'])
            if not color_obj:
                self.stdout.write(self.style.WARNING(
                    f'    Color not found in DB: "{cd["color_name"]}"'
                ))
                continue
            FAColorEvaluation.objects.create(
                evaluation=primary_eval,
                color=color_obj,
                rating=cd['rating'],
                comment=cd['comment'],
            )
            stats['color_evals'] += 1

        # Denormalise onto FA
        fa.primary_inspector = primary_user
        fa.primary_review_date = submitted_at.date() if isinstance(submitted_at, datetime) else submitted_at
        fa.primary_comments = eval_data['primary_comments']
        fa.primary_pattern_execution = eval_data['pattern_execution']
        fa.primary_scale = eval_data['scale']
        fa.primary_spectral_reflectance = eval_data['spectral_reflectance']

        # ---- Final evaluation (only if Crye actually reviewed) ----
        if eval_data['final_result'] in ('pass', 'fail'):
            final_eval = FAEvaluation.objects.create(
                fa=fa,
                stage='final',
                attempt_number=1,
                inspector=final_user,
                pattern_execution=eval_data['pattern_execution'],
                pattern_execution_comment=eval_data['pattern_execution_comment'],
                scale=eval_data['scale'],
                scale_comment=eval_data['scale_comment'],
                spectral_reflectance=eval_data['spectral_reflectance'],
                spectral_reflectance_comment=eval_data['spectral_reflectance_comment'],
                comments=eval_data['final_comments'],
                is_submitted=True,
                submitted_at=submitted_at,
            )
            FAEvaluation.objects.filter(pk=final_eval.pk).update(
                evaluation_date=submitted_at
            )
            stats['final_evals'] += 1

            for cd in eval_data['colors']:
                color_obj = variant_colors.get(cd['color_name'])
                if not color_obj:
                    continue
                FAColorEvaluation.objects.create(
                    evaluation=final_eval,
                    color=color_obj,
                    rating=cd['rating'],
                    comment=cd['comment'],
                )
                stats['color_evals'] += 1

            fa.final_inspector = final_user
            fa.final_review_date = submitted_at.date() if isinstance(submitted_at, datetime) else submitted_at
            fa.final_comments = eval_data['final_comments']
            fa.final_pattern_execution = eval_data['pattern_execution']
            fa.final_scale = eval_data['scale']
            fa.final_spectral_reflectance = eval_data['spectral_reflectance']

        fa.save()
