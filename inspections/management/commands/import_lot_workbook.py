"""
Import Lot Acceptance data from a partner Excel workbook (.xlsx).

Usage:
    python manage.py import_lot_workbook "sheetData/Milliken - Lot Workbook.xlsx" --partner-code MIL
    python manage.py import_lot_workbook "sheetData/Milliken - Lot Workbook.xlsx" --partner-code MIL --dry-run
"""
import re
import openpyxl
from datetime import date, datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.contrib.auth.models import User
from django.utils import timezone

from accounts.models import PartnerCompany
from inspections.models import (
    FirstArticleInspection, LotAcceptance, LotEvaluation,
    LotSampleEvaluation, LotSampleColorEvaluation,
    SHADE_RATING_CHOICES,
)
from core.models import CamouflageType

from inspections.management.commands.import_workbook import (
    to_str, to_date, to_datetime_aware, is_historic_value,
    parse_rating_code, parse_pass_fail, map_variant_name,
    parse_submitter_name,
)

HISTORIC_SENTINEL = 'N/A - Historic First Article'
PLACEHOLDER_DATE = date(2000, 1, 1)

SKIP_SHEETS = frozenset({'Data', 'LotSummary'})


def normalize_sheet_name(s):
    """Strip special chars and truncate to 31 (Excel sheet name limit)."""
    return re.sub(r'[/\\()]', '', s).strip()[:31]


def find_lot_eval_sheet(generated_name, sheet_names):
    """Match a LotSummary generated sheet name to an actual workbook sheet."""
    if not generated_name:
        return None
    g = str(generated_name).strip()
    if g in sheet_names:
        return g
    gn = normalize_sheet_name(g)
    for name in sheet_names:
        nn = normalize_sheet_name(name)
        if gn.startswith(nn) or nn.startswith(gn):
            return name
    return None


def build_sheet_lot_index(wb, skip_sheets):
    """
    Build a lookup { lot_number_str: sheet_name } by reading each eval
    sheet's header (row 3, col 8 = Lot Number/Identifier).
    Handles cases where Excel renames sheets to Sheet5, Sheet6, etc.

    Also indexes by significant numeric substrings (>= 5 digits) to handle
    cases like header "MT-2403228/4000027976" matching lot "PO 4000027976".
    """
    index = {}
    for name in wb.sheetnames:
        if name in skip_sheets:
            continue
        ws = wb[name]
        lot_val = ws.cell(row=3, column=8).value
        if lot_val is not None:
            lot_str = to_str(lot_val)
            if lot_str:
                index[lot_str] = name
                for num in re.findall(r'\d{5,}', lot_str):
                    if num not in index:
                        index[num] = name
    return index


def parse_lot_eval_sheet(ws, num_samples):
    """
    Parse a lot evaluation sheet.

    Each sample block is 10 rows (7 shade colors + 3 criteria) starting at row 7.
    After all samples: "1947 - LOT EVALUATION" header + result row.

    Returns dict with 'samples' list and 'overall_result'/'overall_comments'.
    """
    samples = []
    for s_idx in range(num_samples):
        start_row = 7 + (s_idx * 10)
        sample_id = to_str(ws.cell(row=start_row, column=1).value)

        colors = []
        for c_offset in range(7):
            r = start_row + c_offset
            color_name = ws.cell(row=r, column=3).value
            if not color_name or str(color_name).strip().upper() == 'N/A':
                continue
            colors.append({
                'color_name': str(color_name).strip(),
                'rating': parse_rating_code(ws.cell(row=r, column=4).value),
                'comment': to_str(ws.cell(row=r, column=5).value, 200),
            })

        pe_row = start_row + 7
        sc_row = start_row + 8
        sr_row = start_row + 9

        samples.append({
            'sample_number': s_idx + 1,
            'sample_id': sample_id,
            'colors': colors,
            'pattern_execution': parse_pass_fail(ws.cell(row=pe_row, column=4).value),
            'scale': parse_pass_fail(ws.cell(row=sc_row, column=4).value),
            'spectral_reflectance': parse_pass_fail(ws.cell(row=sr_row, column=4).value),
        })

    eval_header_row = 7 + (num_samples * 10)
    eval_result_row = eval_header_row + 1

    overall_raw = ws.cell(row=eval_result_row, column=1).value
    overall_result = parse_pass_fail(overall_raw) if overall_raw else ''
    overall_comments = to_str(ws.cell(row=eval_result_row, column=4).value)

    return {
        'samples': samples,
        'overall_result': overall_result,
        'overall_comments': overall_comments,
    }


class Command(BaseCommand):
    help = 'Import Lot Acceptance data from a partner Excel workbook'

    def add_arguments(self, parser):
        parser.add_argument('workbook', help='Path to the Excel workbook')
        parser.add_argument('--partner-code', required=True, help='Company code (e.g. MIL)')
        parser.add_argument('--dry-run', action='store_true', help='Preview without saving')

    def handle(self, *args, **options):
        wb_path = options['workbook']
        partner_code = options['partner_code'].upper()
        dry_run = options['dry_run']

        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN — no changes will be made\n'))

        try:
            wb = openpyxl.load_workbook(wb_path, data_only=True)
        except Exception as e:
            raise CommandError(f'Cannot open workbook: {e}')

        if 'LotSummary' not in wb.sheetnames:
            raise CommandError('Workbook missing LotSummary sheet')

        eval_sheet_names = [n for n in wb.sheetnames if n not in SKIP_SHEETS]
        self.stdout.write(f'Evaluation sheets available: {len(eval_sheet_names)}')

        # Index sheets by the lot number in their header (handles Sheet5/6 renames)
        sheet_by_lot_number = build_sheet_lot_index(wb, SKIP_SHEETS)

        # Lookup partner company
        company = None
        if not dry_run:
            try:
                company = PartnerCompany.objects.get(code=partner_code)
            except PartnerCompany.DoesNotExist:
                raise CommandError(
                    f'Partner "{partner_code}" not found. '
                    f'Import FAs first with import_workbook to create the partner.'
                )
            self.stdout.write(self.style.SUCCESS(
                f'Using partner: {company.name} ({company.code})'
            ))
        else:
            self.stdout.write(f'[DRY RUN] Partner code: {partner_code}')

        # Primary inspector (lots only have single-stage review)
        primary_user = None
        if not dry_run:
            try:
                primary_user = User.objects.get(username='primary_inspector')
            except User.DoesNotExist:
                raise CommandError('User "primary_inspector" not found.')

        # Camo type cache for color lookups
        camo_cache = {}
        if not dry_run:
            for ct in CamouflageType.objects.prefetch_related('colors'):
                camo_cache[ct.camouflage_name] = {
                    'obj': ct,
                    'colors': {c.color_name: c for c in ct.colors.all()},
                }
            if not camo_cache:
                raise CommandError('No CamouflageType records. Run load_initial_data first.')

        # Build FA lookup caches for this company
        # fa_by_lot: { fa_lot_number: FA }
        # fa_by_style_variant: { (fabric_style, variant_name): FA }
        # fa_by_style: { fabric_style: FA } (first match only, fallback)
        fa_by_lot = {}
        fa_by_style_variant = {}
        fa_by_style = {}
        if not dry_run:
            for fa in FirstArticleInspection.objects.select_related(
                'multicam_variant'
            ).filter(company=company):
                if fa.fa_lot_number:
                    fa_by_lot[fa.fa_lot_number] = fa
                style = fa.fabric_style.strip()
                variant = fa.multicam_variant.camouflage_name if fa.multicam_variant else ''
                fa_by_style_variant[(style, variant)] = fa
                if style not in fa_by_style:
                    fa_by_style[style] = fa

        stats = dict(
            lots_created=0, evals_created=0, sample_evals=0,
            color_evals=0, errors=0, skipped=0,
            fa_linked_by_lot=0, fa_linked_by_style=0, fa_not_found=0,
        )

        ws = wb['LotSummary']
        row_idx = 3
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            row_idx += 1
            status_raw = row[0]
            if not status_raw or str(status_raw).strip() == '':
                continue

            status = str(status_raw).strip().lower()
            if status == 'pending':
                stats['skipped'] += 1
                continue
            if status not in ('approved', 'rejected'):
                self.stdout.write(self.style.WARNING(
                    f'  Row {row_idx}: Unknown status "{status_raw}", skipping'
                ))
                stats['skipped'] += 1
                continue

            fabric_style = to_str(row[1], 200)
            shade_standard = str(row[3] or '').strip().lower()
            if shade_standard not in ('alpha', 'beta'):
                shade_standard = 'alpha'

            shade_std_num = '' if is_historic_value(row[4]) else to_str(row[4], 20)

            spectral_raw = str(row[5] or '').strip()
            spectral_lower = spectral_raw.lower()
            if spectral_lower in ('alpha', 'beta', 'swir'):
                spectral = spectral_lower
            elif 'visible' in spectral_lower:
                spectral = 'Visible Spectrum Only'
            else:
                spectral = 'alpha'

            orig_fa_lot_raw = row[6]
            orig_fa_lot_str = to_str(orig_fa_lot_raw, 50)
            is_na_fa = is_historic_value(orig_fa_lot_raw)

            lot_lot_number = to_str(row[7], 50)
            if not lot_lot_number:
                lot_lot_number = f'LOT-UNKNOWN-{row_idx}'

            yards_raw = row[8]
            if yards_raw and yards_raw != '':
                try:
                    yards = max(1, int(float(str(yards_raw))))
                except (ValueError, TypeError):
                    yards = 1
            else:
                yards = 1

            num_samples_raw = row[9]
            try:
                num_samples = int(float(str(num_samples_raw)))
            except (ValueError, TypeError):
                num_samples = 2

            date_of_printing = to_date(row[10]) or PLACEHOLDER_DATE
            date_shipped = to_date(row[11])
            tracking = to_str(row[12], 50)
            individual_samples = to_str(row[13]).replace('\n', ', ')
            first_name, last_name = parse_submitter_name(row[14])
            sheet_name = to_str(row[16], 250)

            # Extract variant hint from fabric style (e.g. " - MC - " or " - MCTropic - ")
            variant_hint = None
            variant_match = re.search(r'\s*-\s*(MC\w*)\s*-', fabric_style)
            if variant_match:
                variant_hint = map_variant_name(variant_match.group(1).replace('MC', 'MultiCam ').strip())
                if variant_hint == 'Multicam':
                    variant_hint = 'Multicam'

            # --- Find parent FA ---
            parent_fa = None
            fa_match_method = None
            if not dry_run:
                if not is_na_fa and orig_fa_lot_str:
                    parent_fa = fa_by_lot.get(orig_fa_lot_str)
                    if parent_fa:
                        fa_match_method = 'lot_number'

                if not parent_fa:
                    # Strip variant/lot suffixes to get base fabric style
                    stripped = re.sub(r'\s*-\s*MC\w*\s*-.*$', '', fabric_style).strip()

                    # Try (style, variant) match first for precision
                    if variant_hint and (stripped, variant_hint) in fa_by_style_variant:
                        parent_fa = fa_by_style_variant[(stripped, variant_hint)]
                        fa_match_method = 'fabric_style+variant'
                    elif stripped in fa_by_style:
                        parent_fa = fa_by_style[stripped]
                        fa_match_method = 'fabric_style'
                    else:
                        # Try the raw fabric_style as-is
                        parent_fa = fa_by_style.get(fabric_style.strip())
                        if parent_fa:
                            fa_match_method = 'fabric_style'

                if not parent_fa:
                    self.stdout.write(self.style.ERROR(
                        f'  Row {row_idx}: No parent FA found for '
                        f'fabric="{fabric_style}", orig_fa_lot="{orig_fa_lot_str}". Skipping.'
                    ))
                    stats['fa_not_found'] += 1
                    stats['errors'] += 1
                    continue

            # Skip if already imported (for re-runs)
            if not dry_run:
                existing = LotAcceptance.objects.filter(
                    company=company, lot_lot_number=lot_lot_number
                ).exists()
                if existing:
                    stats['skipped'] += 1
                    continue

            # Find eval sheet: try name match, verify lot# in header, fall back to index
            eval_sheet_name = find_lot_eval_sheet(sheet_name, eval_sheet_names)
            if eval_sheet_name:
                # Verify the sheet actually contains this lot's data
                header_lot = to_str(wb[eval_sheet_name].cell(row=3, column=8).value)
                if header_lot and lot_lot_number and header_lot != lot_lot_number:
                    eval_sheet_name = None
            if not eval_sheet_name:
                eval_sheet_name = sheet_by_lot_number.get(lot_lot_number)
            if not eval_sheet_name:
                for num in re.findall(r'\d{5,}', lot_lot_number):
                    if num in sheet_by_lot_number:
                        eval_sheet_name = sheet_by_lot_number[num]
                        break

            # --- Dry-run output ---
            if dry_run:
                fa_info = f'orig_fa={orig_fa_lot_str}' if not is_na_fa else 'orig_fa=N/A (match by style)'
                self.stdout.write(
                    f'  Row {row_idx}: {status.upper()} | {fabric_style} | '
                    f'Lot: {lot_lot_number} | {num_samples} samples | {fa_info}'
                )
                if eval_sheet_name:
                    self.stdout.write(f'    -> Eval sheet: {eval_sheet_name}')
                else:
                    self.stdout.write(self.style.WARNING(
                        f'    -> No eval sheet for "{sheet_name}"'
                    ))
                stats['lots_created'] += 1
                continue

            # --- Real import ---
            try:
                with transaction.atomic():
                    original_fa_lot = parent_fa.fa_lot_number if parent_fa else orig_fa_lot_str

                    lot = LotAcceptance(
                        company=company,
                        vendor=None,
                        status=status,
                        fabric_style=fabric_style,
                        shade_standard=shade_standard,
                        shade_standard_number=shade_std_num,
                        spectral_reflectance_requirement=spectral,
                        original_fa_lot_number=original_fa_lot,
                        lot_lot_number=lot_lot_number,
                        number_of_yards_printed=yards,
                        number_of_samples=num_samples,
                        individual_sample_numbers=individual_samples or '',
                        date_of_printing=date_of_printing,
                        date_shipped=date_shipped,
                        tracking_number=tracking,
                        submitter_first_name=first_name,
                        submitter_last_name=last_name,
                        original_fa=parent_fa,
                        submitted=True,
                        sheet_name_generated=sheet_name,
                    )
                    lot.save()

                    # Override fields that save() auto-calculated
                    update_fields = {
                        'number_of_samples': num_samples,
                        'lot_lot_number': lot_lot_number,
                    }
                    if individual_samples:
                        update_fields['individual_sample_numbers'] = individual_samples
                    if date_of_printing != PLACEHOLDER_DATE:
                        submission_dt = to_datetime_aware(date_of_printing)
                        if submission_dt:
                            update_fields['submission_date'] = submission_dt

                    LotAcceptance.objects.filter(pk=lot.pk).update(**update_fields)

                    match_info = f'(FA matched by {fa_match_method}: {parent_fa.fai_id})'
                    self.stdout.write(
                        f'  {lot.lot_id}: {fabric_style} | Lot: {lot_lot_number} '
                        f'({status}) {match_info}'
                    )
                    stats['lots_created'] += 1
                    if fa_match_method == 'lot_number':
                        stats['fa_linked_by_lot'] += 1
                    else:
                        stats['fa_linked_by_style'] += 1

                    # Parse evaluation sheet and create evaluation records
                    if eval_sheet_name:
                        eval_ws = wb[eval_sheet_name]
                        eval_data = parse_lot_eval_sheet(eval_ws, num_samples)
                        self._create_lot_evaluation(
                            lot, eval_data, parent_fa, camo_cache,
                            primary_user, date_of_printing, stats,
                        )
                    else:
                        self.stdout.write(self.style.WARNING(
                            f'    No eval sheet for "{sheet_name}"'
                        ))

            except Exception as e:
                stats['errors'] += 1
                self.stdout.write(self.style.ERROR(
                    f'  Row {row_idx}: Error — {e}'
                ))
                import traceback
                traceback.print_exc()

        # Summary
        self.stdout.write('\n' + '=' * 60)
        label = 'DRY RUN SUMMARY' if dry_run else 'IMPORT SUMMARY'
        self.stdout.write(self.style.SUCCESS(f'{label}:'))
        self.stdout.write(f'  Lots created:           {stats["lots_created"]}')
        self.stdout.write(f'    FA linked by lot #:   {stats["fa_linked_by_lot"]}')
        self.stdout.write(f'    FA linked by style:   {stats["fa_linked_by_style"]}')
        self.stdout.write(f'    FA not found:         {stats["fa_not_found"]}')
        self.stdout.write(f'  Lot evaluations:        {stats["evals_created"]}')
        self.stdout.write(f'  Sample evaluations:     {stats["sample_evals"]}')
        self.stdout.write(f'  Color evaluations:      {stats["color_evals"]}')
        self.stdout.write(f'  Errors:                 {stats["errors"]}')
        self.stdout.write(f'  Skipped (Pending):      {stats["skipped"]}')
        self.stdout.write('=' * 60)
        wb.close()

    def _create_lot_evaluation(self, lot, eval_data, parent_fa, camo_cache,
                               primary_user, date_of_printing, stats):
        """Create LotEvaluation, LotSampleEvaluations, and LotSampleColorEvaluations."""
        submitted_at = to_datetime_aware(date_of_printing) or timezone.now()

        variant_name = map_variant_name(
            parent_fa.multicam_variant.camouflage_name
            if parent_fa.multicam_variant else None
        )
        variant_colors = {}
        if variant_name and variant_name in camo_cache:
            variant_colors = camo_cache[variant_name]['colors']
        # Eval sheets often use standard Multicam color names regardless of variant;
        # fall back to Multicam colors so shade ratings aren't lost
        mc_colors = camo_cache.get('Multicam', {}).get('colors', {})

        lot_eval = LotEvaluation.objects.create(
            lot=lot,
            inspector=primary_user,
            comments=eval_data['overall_comments'],
            is_submitted=True,
            submitted_at=submitted_at,
        )
        LotEvaluation.objects.filter(pk=lot_eval.pk).update(
            evaluation_date=submitted_at
        )
        stats['evals_created'] += 1

        for sample_data in eval_data['samples']:
            sample_eval = LotSampleEvaluation.objects.create(
                lot_evaluation=lot_eval,
                sample_number=sample_data['sample_number'],
                sample_id=sample_data['sample_id'],
                pattern_execution=sample_data['pattern_execution'],
                scale=sample_data['scale'],
                spectral_reflectance=sample_data['spectral_reflectance'],
            )
            stats['sample_evals'] += 1

            for cd in sample_data['colors']:
                color_obj = variant_colors.get(cd['color_name'])
                if not color_obj:
                    color_obj = mc_colors.get(cd['color_name'])
                if not color_obj:
                    self.stdout.write(self.style.WARNING(
                        f'    Color not found: "{cd["color_name"]}"'
                    ))
                    continue
                LotSampleColorEvaluation.objects.create(
                    sample_evaluation=sample_eval,
                    color=color_obj,
                    rating=cd['rating'],
                    comment=cd['comment'],
                )
                stats['color_evals'] += 1

        # Update denormalized fields on LotAcceptance
        lot.inspector = primary_user
        lot.review_date = submitted_at.date() if hasattr(submitted_at, 'date') else submitted_at
        lot.inspector_comments = eval_data['overall_comments']
        lot.save(update_fields=[
            'inspector', 'review_date', 'inspector_comments',
        ])

        # Override individual_sample_numbers with actual IDs from eval sheet
        actual_ids = [s['sample_id'] for s in eval_data['samples'] if s['sample_id']]
        if actual_ids:
            LotAcceptance.objects.filter(pk=lot.pk).update(
                individual_sample_numbers=', '.join(actual_ids)
            )
